from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import os
import cv2
import face_recognition
import numpy as np
from datetime import datetime, timedelta
import csv
import base64
import pickle
import threading
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change this in production
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Session configuration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)  # 8 hour session
app.config['SESSION_COOKIE_SECURE'] = False  # Set to True in production with HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

db = SQLAlchemy(app)

def validate_session():
    """Validate and refresh session if needed"""
    if 'teacher_id' in session and 'login_time' in session:
        try:
            login_time = datetime.fromisoformat(session['login_time'])
            if datetime.now() - login_time > timedelta(hours=8):
                # Session expired
                session.clear()
                return False
            
            # Refresh session if it's getting old (refresh every 2 hours)
            if datetime.now() - login_time > timedelta(hours=2):
                session['login_time'] = datetime.now().isoformat()
                session.modified = True
            
            return True
        except:
            session.clear()
            return False
    return False

# Database Models
class Teacher(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    roll_number = db.Column(db.String(20), unique=True, nullable=False)
    class_name = db.Column(db.String(50), nullable=False)
    
    def to_dict(self):
        """Convert Student object to dictionary for JSON serialization"""
        return {
            'id': self.id,
            'name': self.name,
            'roll_number': self.roll_number,
            'class_name': self.class_name
        }

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    time = db.Column(db.Time, nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('teacher.id'), nullable=False)
    student = db.relationship('Student', backref='attendance_records')

# Load known face encodings with improved accuracy
def load_known_faces():
    dataset_path = "dataset"
    known_encodings = []
    known_names = []
    student_encodings = {}  # Store multiple encodings per student
    
    for student_name in os.listdir(dataset_path):
        student_folder = os.path.join(dataset_path, student_name)
        
        if not os.path.isdir(student_folder):
            continue
        
        student_encodings[student_name] = []
        
        loaded_count = 0
        skipped_count = 0
        # Process all images for each student
        for file in os.listdir(student_folder):
            if file.endswith((".jpg", ".png")):
                img_path = os.path.join(student_folder, file)
                img = cv2.imread(img_path)
                if img is None:
                    continue
                
                # Resize large images for better processing
                height, width = img.shape[:2]
                if max(height, width) > 1280:
                    scale = 1280 / max(height, width)
                    new_width = int(width * scale)
                    new_height = int(height * scale)
                    img = cv2.resize(img, (new_width, new_height))
                
                rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

                # Try multiple face detection strategies
                face_locations = face_recognition.face_locations(rgb_img, model="hog", number_of_times_to_upsample=1)
                if len(face_locations) == 0:
                    face_locations = face_recognition.face_locations(rgb_img, model="hog", number_of_times_to_upsample=2)
                if len(face_locations) == 0:
                    try:
                        face_locations = face_recognition.face_locations(rgb_img, model="cnn", number_of_times_to_upsample=0)
                    except Exception:
                        face_locations = []

                if len(face_locations) > 0:
                    # If multiple faces, use the largest one
                    if len(face_locations) > 1:
                        def face_area(face_location):
                            top, right, bottom, left = face_location
                            return (bottom - top) * (right - left)
                        face_locations = [max(face_locations, key=face_area)]
                    
                    encodings = face_recognition.face_encodings(rgb_img, face_locations)
                    if len(encodings) > 0:
                        student_encodings[student_name].append(encodings[0])
                        loaded_count += 1
                    else:
                        skipped_count += 1
                else:
                    skipped_count += 1
        print(f"ℹ️  Loaded encodings for {student_name}: {loaded_count} images, skipped {skipped_count}")
    
    # Add all encodings to the main lists
    for student_name, encodings in student_encodings.items():
        if encodings:  # Only add if we have at least one encoding
            for encoding in encodings:
                known_encodings.append(encoding)
                known_names.append(student_name)
    
    print(f"✅ Loaded {len(set(known_names))} students with {len(known_encodings)} total encodings")
    return known_encodings, known_names

# Improved face recognition with confidence scoring
def recognize_face_with_confidence(face_encoding, known_encodings, known_names, confidence_threshold=0.6):
    """
    Recognize a face with confidence scoring
    Returns (name, confidence) or (None, 0) if below threshold
    """
    if len(known_encodings) == 0:
        return None, 0
    
    # Calculate face distances
    face_distances = face_recognition.face_distance(known_encodings, face_encoding)
    
    # Find the best match
    best_match_index = np.argmin(face_distances)
    best_distance = face_distances[best_match_index]
    
    # Convert distance to confidence (0-1 scale, higher is better)
    # face_recognition uses 0.6 as default threshold, so we'll use that as our baseline
    confidence = max(0, 1 - (best_distance / 0.6))
    
    if confidence >= confidence_threshold:
        return known_names[best_match_index], confidence
    else:
        return None, confidence

# Cache known faces to avoid re-loading on every request
KNOWN_FACE_DATA = {
    'encodings': [],
    'names': [],
    'dataset_mtime': 0.0,
    'encodings_pkl': 'encodings.pkl',
    'is_loading': False
}
_LOAD_LOCK = threading.Lock()

def get_dataset_mtime(path="dataset"):
    try:
        latest = os.path.getmtime(path)
        for root, dirs, files in os.walk(path):
            for name in files:
                try:
                    latest = max(latest, os.path.getmtime(os.path.join(root, name)))
                except Exception:
                    pass
        return latest
    except Exception:
        return 0.0

def ensure_known_faces_loaded():
    current_mtime = get_dataset_mtime()
    need_reload = (
        not KNOWN_FACE_DATA['encodings'] or not KNOWN_FACE_DATA['names'] or
        current_mtime > KNOWN_FACE_DATA['dataset_mtime']
    )

    if not need_reload:
        return KNOWN_FACE_DATA['encodings'], KNOWN_FACE_DATA['names']

    with _LOAD_LOCK:
        # Re-check inside lock
        current_mtime = get_dataset_mtime()
        need_reload = (
            not KNOWN_FACE_DATA['encodings'] or not KNOWN_FACE_DATA['names'] or
            current_mtime > KNOWN_FACE_DATA['dataset_mtime']
        )
        if not need_reload:
            return KNOWN_FACE_DATA['encodings'], KNOWN_FACE_DATA['names']

        KNOWN_FACE_DATA['is_loading'] = True
        try:
            # Try loading from pickle if fresh
            pkl_path = KNOWN_FACE_DATA['encodings_pkl']
            try:
                pkl_mtime = os.path.getmtime(pkl_path)
            except Exception:
                pkl_mtime = 0.0

            if pkl_mtime >= current_mtime and os.path.exists(pkl_path):
                with open(pkl_path, 'rb') as f:
                    data = pickle.load(f)
                encs = data.get('encodings', [])
                names = data.get('names', [])
                print(f"✅ Loaded encodings from {pkl_path} ({len(set(names))} students, {len(encs)} encodings)")
            else:
                encs, names = load_known_faces()
                # Save to pickle for future fast starts
                try:
                    with open(pkl_path, 'wb') as f:
                        pickle.dump({'encodings': encs, 'names': names}, f)
                    print(f"💾 Saved encodings to {pkl_path}")
                except Exception as e:
                    print(f"⚠️  Could not save encodings to pickle: {e}")

            KNOWN_FACE_DATA['encodings'] = encs
            KNOWN_FACE_DATA['names'] = names
            KNOWN_FACE_DATA['dataset_mtime'] = current_mtime
        finally:
            KNOWN_FACE_DATA['is_loading'] = False

    return KNOWN_FACE_DATA['encodings'], KNOWN_FACE_DATA['names']

@app.route('/api/recognize', methods=['POST'])
def api_recognize():
    if not validate_session():
        return jsonify({'error': 'Not authenticated'}), 401
    try:
        data = request.get_json(silent=True) or {}
        image_data_url = data.get('image')
        if not image_data_url or not isinstance(image_data_url, str):
            return jsonify({'error': 'No image provided'}), 400

        # Strip data URL header if present
        if ',' in image_data_url:
            image_b64 = image_data_url.split(',', 1)[1]
        else:
            image_b64 = image_data_url

        image_bytes = base64.b64decode(image_b64)
        np_arr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        if frame is None:
            return jsonify({'error': 'Invalid image data'}), 400
        h, w = frame.shape[:2]
        print(f"DEBUG: Decoded frame {w}x{h}")

        # If the frame is very small, upscale before detection
        if max(h, w) < 400:
            scale = 400.0 / max(h, w)
            new_w = int(w * scale)
            new_h = int(h * scale)
            frame = cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_CUBIC)
            print(f"DEBUG: Upscaled frame to {new_w}x{new_h}")

        # Lean fast-path detection: assume client already downscaled
        rgb_small = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_small, number_of_times_to_upsample=1, model="hog")
        if not face_locations:
            # One extra upsample pass as fallback
            face_locations = face_recognition.face_locations(rgb_small, number_of_times_to_upsample=2, model="hog")
        if not face_locations:
            # Last resort: try at slightly larger scale
            try:
                bigger = cv2.resize(rgb_small, (0, 0), fx=1.25, fy=1.25)
                face_locations = face_recognition.face_locations(bigger, number_of_times_to_upsample=2, model="hog")
                if face_locations:
                    rgb_small = bigger
            except Exception:
                pass

        # If still nothing, try light enhancement (CLAHE on Y channel)
        if not face_locations:
            try:
                yuv = cv2.cvtColor(frame, cv2.COLOR_BGR2YUV)
                y, u, v = cv2.split(yuv)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                y_eq = clahe.apply(y)
                yuv_eq = cv2.merge((y_eq, u, v))
                bgr_eq = cv2.cvtColor(yuv_eq, cv2.COLOR_YUV2BGR)
                rgb_eq = cv2.cvtColor(bgr_eq, cv2.COLOR_BGR2RGB)
                face_locations = face_recognition.face_locations(rgb_eq, number_of_times_to_upsample=2, model="hog")
                if face_locations:
                    rgb_small = rgb_eq
            except Exception:
                pass

        face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

        known_encodings, known_names = ensure_known_faces_loaded()

        detections = []
        print(f"DEBUG: /api/recognize faces={len(face_locations)}")
        for face_encoding, face_location in zip(face_encodings, face_locations):
            name, confidence = recognize_face_with_confidence(
                face_encoding,
                known_encodings,
                known_names,
                confidence_threshold=0.50
            )
            if name is None:
                name = 'Unknown'
                confidence = float(confidence)
            else:
                # Log best match for debugging
                print(f"DEBUG: matched name={name} conf={confidence:.2f}")
            detections.append({
                'name': name,
                'confidence': float(confidence)
            })

        return jsonify({'success': True, 'detections': detections})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/')
def index():
    if 'teacher_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        teacher = Teacher.query.filter_by(username=username).first()
        
        if teacher and check_password_hash(teacher.password_hash, password):
            # Clear any existing session and create new one
            session.clear()
            session['teacher_id'] = teacher.id
            session['teacher_name'] = teacher.name
            session['login_time'] = datetime.now().isoformat()
            session.modified = True
            
            # Debug information
            print(f"DEBUG: Teacher {teacher.name} (ID: {teacher.id}) logged in successfully")
            print(f"DEBUG: Session created with teacher_id: {session.get('teacher_id')}")
            
            flash(f'Login successful! Welcome, {teacher.name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    if 'teacher_name' in session:
        teacher_name = session['teacher_name']
        session.clear()
        flash(f'Logged out successfully, {teacher_name}!', 'success')
    else:
        session.clear()
        flash('Logged out successfully!', 'success')
    return redirect(url_for('index'))

@app.route('/debug_session')
def debug_session():
    """Debug route to check current session (remove in production)"""
    if 'teacher_id' in session:
        return jsonify({
            'teacher_id': session.get('teacher_id'),
            'teacher_name': session.get('teacher_name'),
            'login_time': session.get('login_time'),
            'session_valid': validate_session()
        })
    else:
        return jsonify({'message': 'No active session'})

@app.route('/force_logout')
def force_logout():
    """Force logout and clear all sessions (for testing)"""
    session.clear()
    flash('All sessions cleared. Please login again.', 'info')
    return redirect(url_for('login'))

@app.route('/list_teachers')
def list_teachers():
    """List all teachers (for debugging - remove in production)"""
    teachers = Teacher.query.all()
    teacher_list = []
    for teacher in teachers:
        teacher_list.append({
            'id': teacher.id,
            'username': teacher.username,
            'name': teacher.name,
            'email': teacher.email
        })
    return jsonify(teacher_list)

@app.route('/dashboard')
def dashboard():
    if not validate_session():
        return redirect(url_for('login'))
    
    # Get today's attendance
    today = datetime.now().date()
    today_attendance = Attendance.query.filter_by(date=today).all()
    
    # Get all students and convert to dictionaries for JSON serialization
    students = Student.query.all()
    students_data = [student.to_dict() for student in students]
    
    return render_template('dashboard.html', 
                         students=students_data, 
                         today_attendance=today_attendance,
                         today=today)

@app.route('/take_attendance')
def take_attendance():
    if not validate_session():
        return redirect(url_for('login'))
    
    students = Student.query.all()
    # Convert Student objects to dictionaries for JSON serialization
    students_data = [student.to_dict() for student in students]
    
    return render_template('take_attendance.html', students=students_data)

@app.route('/api/process_attendance', methods=['POST'])
def process_attendance():
    if not validate_session():
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        payload = request.get_json(silent=True) or {}
        # Support both current and legacy keys
        detected_students = payload.get('detected_students') or payload.get('recognized') or payload.get('detections') or []
        if not isinstance(detected_students, list):
            return jsonify({'error': 'Invalid payload: expected list of students'}), 400
        
        # Normalize to list of names (client may send objects)
        normalized_names = []
        for item in detected_students:
            if isinstance(item, dict):
                # Accept {name: str} or full student object
                name = item.get('name') if 'name' in item else None
                if name:
                    normalized_names.append(name)
            elif isinstance(item, str):
                normalized_names.append(item)
        
        if not normalized_names:
            return jsonify({'success': True, 'message': 'No students to mark', 'marked_count': 0})
        today = datetime.now().date()
        current_time = datetime.now().time()
        
        marked_count = 0
        for student_name in normalized_names:
            student = Student.query.filter_by(name=student_name).first()
            if student:
                # Check if attendance already marked
                existing = Attendance.query.filter_by(
                    student_id=student.id, 
                    date=today
                ).first()
                
                if not existing:
                    attendance = Attendance(
                        student_id=student.id,
                        date=today,
                        time=current_time,
                        teacher_id=session['teacher_id']
                    )
                    db.session.add(attendance)
                    marked_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Attendance marked for {marked_count} students',
            'marked_count': marked_count
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/manual_mark', methods=['POST'])
def manual_mark():
    if not validate_session():
        return redirect(url_for('login'))

    try:
        # Prefer form submission from dashboard; fallback to JSON body
        student_id = request.form.get('student_id')
        student_name = request.form.get('student_name')

        if not student_id and not student_name:
            data = request.get_json(silent=True) or {}
            student_id = data.get('student_id')
            student_name = data.get('student_name') or data.get('name')

        student = None
        if student_id:
            try:
                student = Student.query.get(int(student_id))
            except Exception:
                student = None
        if student is None and student_name:
            student = Student.query.filter_by(name=student_name).first()

        if not student:
            if request.is_json:
                return jsonify({'success': False, 'error': 'Student not found'}), 404
            flash('Student not found', 'error')
            return redirect(url_for('dashboard'))

        today = datetime.now().date()
        current_time = datetime.now().time()

        existing = Attendance.query.filter_by(student_id=student.id, date=today).first()
        if existing:
            if request.is_json:
                return jsonify({'success': True, 'message': 'Attendance already marked'}), 200
            flash(f'Attendance already marked for {student.name}', 'info')
            return redirect(url_for('dashboard'))

        attendance = Attendance(
            student_id=student.id,
            date=today,
            time=current_time,
            teacher_id=session['teacher_id']
        )
        db.session.add(attendance)
        db.session.commit()

        if request.is_json:
            return jsonify({'success': True, 'message': f'Marked present: {student.name}'}), 200
        flash(f'Marked present: {student.name}', 'success')
        return redirect(url_for('dashboard'))
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 500
        flash('Failed to mark attendance manually', 'error')
        return redirect(url_for('dashboard'))

@app.route('/api/manual_mark', methods=['POST'])
def api_manual_mark():
    if not validate_session():
        return jsonify({'error': 'Not authenticated'}), 401

    data = request.get_json(silent=True) or {}
    # Delegate to manual_mark logic by simulating JSON path
    with app.test_request_context(json=data):
        return manual_mark()

@app.route('/export_attendance')
def export_attendance():
    if not validate_session():
        return redirect(url_for('login'))
    
    # Get today's attendance
    today = datetime.now().date()
    today_attendance = Attendance.query.filter_by(date=today).all()
    
    # Get all students to show absent ones
    all_students = Student.query.all()
    present_student_ids = {att.student_id for att in today_attendance}
    
    # Create CSV content
    csv_content = "Student Name,Roll Number,Class,Date,Time,Status\n"
    
    # Add present students
    for attendance in today_attendance:
        student = attendance.student
        csv_content += f'"{student.name}","{student.roll_number}","{student.class_name}","{today.strftime("%Y-%m-%d")}","{attendance.time.strftime("%H:%M:%S")}","Present"\n'
    
    # Add absent students
    for student in all_students:
        if student.id not in present_student_ids:
            csv_content += f'"{student.name}","{student.roll_number}","{student.class_name}","{today.strftime("%Y-%m-%d")}","","Absent"\n'
    
    # Create response with CSV file
    from flask import Response
    response = Response(csv_content, mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename=attendance_{today.strftime("%Y-%m-%d")}.csv'
    
    return response

@app.route('/export_attendance_excel')
def export_attendance_excel():
    if not validate_session():
        return redirect(url_for('login'))
    try:
        # Lazy import to keep startup fast
        import openpyxl
        from openpyxl.utils import get_column_letter

        today = datetime.now().date()
        today_attendance = Attendance.query.filter_by(date=today).all()
        all_students = Student.query.all()
        present_student_ids = {att.student_id for att in today_attendance}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance {today.strftime('%Y-%m-%d')}"

        headers = ["Student Name", "Roll Number", "Class", "Date", "Time", "Status"]
        ws.append(headers)

        for attendance in today_attendance:
            s = attendance.student
            ws.append([s.name, s.roll_number, s.class_name, today.strftime('%Y-%m-%d'), attendance.time.strftime('%H:%M:%S'), 'Present'])

        for s in all_students:
            if s.id not in present_student_ids:
                ws.append([s.name, s.roll_number, s.class_name, today.strftime('%Y-%m-%d'), '', 'Absent'])

        # Auto width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        from flask import send_file
        filename = f"attendance_{today.strftime('%Y-%m-%d')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError as e:
        print(f"ERROR: openpyxl import failed: {e}")
        return jsonify({'error': 'Excel support not installed. Run: pip install openpyxl'}), 500
    except Exception as e:
        print(f"ERROR: export_attendance_excel failed: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export_attendance_range')
def export_attendance_range():
    if not validate_session():
        return redirect(url_for('login'))
    
    # Get date range from query parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({'error': 'Start date and end date are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    
    # Get attendance for date range
    attendance_records = Attendance.query.filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).order_by(Attendance.date, Attendance.time).all()
    
    # Get all students
    all_students = Student.query.all()
    
    # Create CSV content
    csv_content = "Student Name,Roll Number,Class,Date,Time,Status\n"
    
    # Group attendance by date
    attendance_by_date = {}
    for att in attendance_records:
        date_str = att.date.strftime('%Y-%m-%d')
        if date_str not in attendance_by_date:
            attendance_by_date[date_str] = []
        attendance_by_date[date_str].append(att)
    
    # Generate CSV for each date in range
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        present_student_ids = {att.student_id for att in attendance_by_date.get(date_str, [])}
        
        # Add present students for this date
        for attendance in attendance_by_date.get(date_str, []):
            student = attendance.student
            csv_content += f'"{student.name}","{student.roll_number}","{student.class_name}","{date_str}","{attendance.time.strftime("%H:%M:%S")}","Present"\n'
        
        # Add absent students for this date
        for student in all_students:
            if student.id not in present_student_ids:
                csv_content += f'"{student.name}","{student.roll_number}","{student.class_name}","{date_str}","","Absent"\n'
        
        current_date = current_date + timedelta(days=1)
    
    # Create response with CSV file
    from flask import Response
    filename = f'attendance_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.csv'
    response = Response(csv_content, mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@app.route('/export_attendance_range_excel')
def export_attendance_range_excel():
    if not validate_session():
        return redirect(url_for('login'))
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        if not start_date or not end_date:
            return jsonify({'error': 'Start date and end date are required'}), 400

        try:
            start_date_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        import openpyxl
        from openpyxl.utils import get_column_letter

        attendance_records = Attendance.query.filter(
            Attendance.date >= start_date_dt,
            Attendance.date <= end_date_dt
        ).order_by(Attendance.date, Attendance.time).all()

        all_students = Student.query.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{start_date_dt.strftime('%Y-%m-%d')} to {end_date_dt.strftime('%Y-%m-%d')}"

        headers = ["Student Name", "Roll Number", "Class", "Date", "Time", "Status"]
        ws.append(headers)

        attendance_by_date = {}
        for att in attendance_records:
            date_str = att.date.strftime('%Y-%m-%d')
            attendance_by_date.setdefault(date_str, []).append(att)

        current_date = start_date_dt
        while current_date <= end_date_dt:
            date_str = current_date.strftime('%Y-%m-%d')
            present_ids = {att.student_id for att in attendance_by_date.get(date_str, [])}

            for att in attendance_by_date.get(date_str, []):
                s = att.student
                ws.append([s.name, s.roll_number, s.class_name, date_str, att.time.strftime('%H:%M:%S'), 'Present'])

            for s in all_students:
                if s.id not in present_ids:
                    ws.append([s.name, s.roll_number, s.class_name, date_str, '', 'Absent'])

            current_date = current_date + timedelta(days=1)

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        from flask import send_file
        filename = f"attendance_{start_date_dt.strftime('%Y-%m-%d')}_to_{end_date_dt.strftime('%Y-%m-%d')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError as e:
        print(f"ERROR: openpyxl import failed: {e}")
        return jsonify({'error': 'Excel support not installed. Run: pip install openpyxl'}), 500
    except Exception as e:
        print(f"ERROR: export_attendance_range_excel failed: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export_attendance_range_graph')
def export_attendance_range_graph():
    if not validate_session():
        return redirect(url_for('login'))

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if not start_date or not end_date:
        return jsonify({'error': 'Start date and end date are required'}), 400

    try:
        start_date_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    try:
        import matplotlib
        matplotlib.use('Agg')  # headless backend
        import matplotlib.pyplot as plt

        # Query attendance grouped by date
        attendance_records = Attendance.query.filter(
            Attendance.date >= start_date_dt,
            Attendance.date <= end_date_dt
        ).all()

        # Build date range and counts
        date_cursor = start_date_dt
        dates = []
        counts = []
        while date_cursor <= end_date_dt:
            dates.append(date_cursor.strftime('%Y-%m-%d'))
            count_for_day = sum(1 for att in attendance_records if att.date == date_cursor)
            counts.append(count_for_day)
            date_cursor = date_cursor + timedelta(days=1)

        # Plot
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(dates, counts, marker='o', linewidth=2, color='#0d6efd')
        ax.set_title(f"Attendance count per day ({start_date} to {end_date})")
        ax.set_xlabel('Date')
        ax.set_ylabel('Present Count')
        ax.grid(True, linestyle='--', alpha=0.4)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()

        bio = BytesIO()
        plt.savefig(bio, format='png')
        plt.close(fig)
        bio.seek(0)
        from flask import send_file
        filename = f"attendance_graph_{start_date}_to_{end_date}.png"
        return send_file(bio, as_attachment=True, download_name=filename, mimetype='image/png')
    except ImportError as e:
        print(f"ERROR: matplotlib import failed: {e}")
        return jsonify({'error': 'Graph support not installed. Run: pip install matplotlib'}), 500
    except Exception as e:
        print(f"ERROR: export_attendance_range_graph failed: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/register_teacher', methods=['GET', 'POST'])
def register_teacher():
    # Clear any existing session when registering
    session.clear()
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        name = request.form['name']
        email = request.form['email']
        
        # Check if username already exists
        if Teacher.query.filter_by(username=username).first():
            flash('Username already exists', 'error')
            return render_template('register_teacher.html')
        
        # Check if email already exists
        if Teacher.query.filter_by(email=email).first():
            flash('Email already exists', 'error')
            return render_template('register_teacher.html')
        
        # Create new teacher
        teacher = Teacher(
            username=username,
            password_hash=generate_password_hash(password),
            name=name,
            email=email
        )
        
        db.session.add(teacher)
        db.session.commit()
        
        # Verify the teacher was created
        created_teacher = Teacher.query.filter_by(username=username).first()
        if created_teacher:
            flash(f'Teacher "{name}" registered successfully with username "{username}"! Please login.', 'success')
        else:
            flash('Registration completed but verification failed. Please try logging in.', 'warning')
        
        return redirect(url_for('login'))
    
    return render_template('register_teacher.html')

@app.route('/add_student', methods=['GET', 'POST'])
def add_student():
    if not validate_session():
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        name = request.form['name']
        roll_number = request.form['roll_number']
        class_name = request.form['class_name']
        
        # Check if roll number already exists
        if Student.query.filter_by(roll_number=roll_number).first():
            flash('Roll number already exists', 'error')
            return render_template('add_student.html')
        
        # Create new student
        student = Student(
            name=name,
            roll_number=roll_number,
            class_name=class_name
        )
        
        db.session.add(student)
        db.session.commit()
        
        flash('Student added successfully!', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('add_student.html')

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Create a default teacher if none exists
        if not Teacher.query.first():
            default_teacher = Teacher(
                username='admin',
                password_hash=generate_password_hash('admin123'),
                name='Administrator',
                email='admin@school.com'
            )
            db.session.add(default_teacher)
            db.session.commit()
            print("Default teacher created: username='admin', password='admin123'")
    
    # Disable reloader/threaded mode to avoid native lib crashes with dlib/OpenCV
    app.run(debug=False, use_reloader=False, threaded=False, host='0.0.0.0', port=5000)
